"""Bulk import 관세청 품목번호별 관세율표 into ``customs_duty_rates``.

Pull only two rate types per HS code:
* ``A``    = 기본관세 (base duty)
* ``FCN1`` = 한-중 FTA (KCFTA) primary-year rate

The ``FCN2`` / ``FCN3`` stages are ignored (they're phased-in tariff
cut schedules for the same HS; we just use stage 1 as the effective
FTA rate). The operator can still override per-product via the edit
form when the staging matters.

Workflow
--------
1. Download from https://www.data.go.kr/data/15051179/fileData.do
   (파일데이터 → 다운로드).
2. ::

     cd backend
     .venv/bin/python -m scripts.import_customs_rates \\
         --file ~/projects/itemSelector/관세청_품목번호별\\ 관세율표_20260211.xlsx
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.db.session import SessionLocal
from app.models import CustomsDutyRate


HEADERS = {
    "품목번호": "hs_code",
    "관세율구분": "kind",
    "관세율": "rate",
    "적용개시일": "start",
    "적용만료일": "end",
}


def _normalise_rate(raw) -> Decimal | None:
    """Accepts ``"8"`` or ``"8.0"`` or ``"0%"`` → Decimal(0.08)."""
    if raw is None or raw == "":
        return None
    s = str(raw).strip().replace("%", "").replace(",", "")
    if not s:
        return None
    try:
        value = Decimal(s)
    except Exception:
        return None
    if value < 0:
        return None
    # File stores percent form (8 = 8%). Normalise to decimal (0.08).
    return (value / Decimal(100)).quantize(Decimal("0.0001"))


def _parse_date(raw) -> date | None:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s, "%Y%m%d").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        return None


def _resolve_columns(header: list) -> dict[str, int]:
    out: dict[str, int] = {}
    for idx, cell in enumerate(header):
        label = str(cell or "").strip()
        if label in HEADERS:
            out[HEADERS[label]] = idx
    missing = {"hs_code", "kind", "rate"} - set(out)
    if missing:
        raise SystemExit(
            f"required columns missing: {missing}\nheader: {header}"
        )
    return out


def import_file(path: Path, *, dry_run: bool = False) -> dict[str, int]:
    wb = load_workbook(path, read_only=True, data_only=True)

    # Pass 1: scan every sheet, collect A and FCN1 rows per HS.
    by_hs: dict[str, dict] = {}
    rows_seen = 0

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        try:
            header = list(next(rows))
        except StopIteration:
            continue
        cols = _resolve_columns(header)

        for row in rows:
            rows_seen += 1
            hs_raw = row[cols["hs_code"]]
            kind_raw = row[cols["kind"]]
            if not hs_raw or not kind_raw:
                continue
            hs = str(hs_raw).strip()
            kind = str(kind_raw).strip()
            if kind not in ("A", "FCN1"):
                continue
            digits = "".join(c for c in hs if c.isdigit())
            if len(digits) != 10:
                continue
            rate = _normalise_rate(row[cols["rate"]])
            if rate is None:
                # 종량세/공란 — skip, will leave as NULL.
                continue
            start = (
                _parse_date(row[cols["start"]]) if "start" in cols else None
            )
            end = _parse_date(row[cols["end"]]) if "end" in cols else None

            bucket = by_hs.setdefault(
                digits,
                {"base_duty_pct": None, "kcfta_duty_pct": None,
                 "effective_start": None, "effective_end": None},
            )
            if kind == "A":
                bucket["base_duty_pct"] = rate
            else:  # FCN1
                bucket["kcfta_duty_pct"] = rate
            # Keep the widest time window across matched rows.
            if start and (bucket["effective_start"] is None or start < bucket["effective_start"]):
                bucket["effective_start"] = start
            if end and (bucket["effective_end"] is None or end > bucket["effective_end"]):
                bucket["effective_end"] = end

    # Pass 2: upsert.
    session = SessionLocal()
    inserted = 0
    skipped_both_null = 0
    try:
        batch: list[dict] = []
        for hs, bucket in by_hs.items():
            if bucket["base_duty_pct"] is None and bucket["kcfta_duty_pct"] is None:
                skipped_both_null += 1
                continue
            batch.append({"hs_code": hs, **bucket})
            if len(batch) >= 500:
                inserted += _flush(session, batch, dry_run=dry_run)
                batch.clear()
        if batch:
            inserted += _flush(session, batch, dry_run=dry_run)
        if not dry_run:
            session.commit()
    finally:
        session.close()

    return {
        "rows_seen": rows_seen,
        "hs_codes_covered": len(by_hs),
        "inserted_or_updated": inserted,
        "skipped_no_rate": skipped_both_null,
    }


def _flush(session, batch: list[dict], *, dry_run: bool) -> int:
    if dry_run:
        return len(batch)
    stmt = pg_insert(CustomsDutyRate).values(batch)
    stmt = stmt.on_conflict_do_update(
        index_elements=["hs_code"],
        set_={
            "base_duty_pct": stmt.excluded.base_duty_pct,
            "kcfta_duty_pct": stmt.excluded.kcfta_duty_pct,
            "effective_start": stmt.excluded.effective_start,
            "effective_end": stmt.excluded.effective_end,
        },
    )
    session.execute(stmt)
    return len(batch)


def _main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args(argv)

    if not args.file.exists():
        raise SystemExit(f"file not found: {args.file}")

    result = import_file(args.file, dry_run=args.dry_run)
    print(f"rows_seen:          {result['rows_seen']:,}")
    print(f"hs_codes_covered:   {result['hs_codes_covered']:,}")
    print(f"inserted_or_updated: {result['inserted_or_updated']:,}")
    print(f"skipped (no rate):  {result['skipped_no_rate']:,}")
    return 0


if __name__ == "__main__":
    sys.exit(_main())
