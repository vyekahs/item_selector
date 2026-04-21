"""Bulk import 관세청 HS부호 from a downloaded Excel file.

The dataset on data.go.kr is published as ``fileData`` (.xlsx), not an
HTTP API. Workflow:

1. Visit https://www.data.go.kr/data/15049722/fileData.do
2. Sign in → "파일데이터 다운로드" → save the .xlsx locally.
3. Run::

       cd backend
       .venv/bin/python -m scripts.import_hs_codes \\
           --file ~/Downloads/HS부호_20260101.xlsx

Re-running is **idempotent** — rows are upserted on ``code``. Only
the 6 and 10 digit rows are kept (matches the ``code_length_6_or_10``
CHECK constraint on ``hs_codes``).

Column mapping
--------------
The published file changes column headers occasionally. We match on
substring rather than exact label so the script survives renames:

* ``HS부호`` / ``HS Code``  → ``code``
* ``한글품목명`` / ``품목명_한글`` → ``name_ko``
* ``영문품목명`` / ``품목명_영문`` → ``name_en``
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.db.session import SessionLocal
from app.models import HsCode


def _normalise(s: str) -> str:
    return "".join(ch for ch in s.lower() if not ch.isspace())


def _resolve_columns(header: list[str]) -> dict[str, int]:
    """Map our logical column → 0-based index in the workbook header."""
    out: dict[str, int] = {}
    for idx, raw in enumerate(header):
        norm = _normalise(str(raw or ""))
        if "code" in out and "name_ko" in out and "name_en" in out:
            break
        if "code" not in out and ("hs부호" in norm or norm == "hscode" or "hs코드" in norm):
            out["code"] = idx
        elif "name_ko" not in out and ("한글품목명" in norm or "품목명한글" in norm):
            out["name_ko"] = idx
        elif "name_en" not in out and ("영문품목명" in norm or "품목명영문" in norm):
            out["name_en"] = idx
    missing = {"code", "name_ko"} - set(out)
    if missing:
        raise SystemExit(
            f"required columns not found in workbook header: {missing}\n"
            f"got header: {header}"
        )
    return out


def _normalise_code(raw: object) -> str | None:
    if raw is None:
        return None
    s = str(raw).strip().replace(".", "").replace("-", "")
    if not s.isdigit():
        return None
    if len(s) not in (6, 10):
        return None
    return s


def import_file(path: Path, *, dry_run: bool = False) -> dict[str, int]:
    if not path.exists():
        raise SystemExit(f"file not found: {path}")
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    header = list(next(rows))
    cols = _resolve_columns(header)

    inserted = 0
    skipped = 0
    seen_codes: set[str] = set()

    session = SessionLocal()
    try:
        batch: list[dict] = []
        for row in rows:
            code = _normalise_code(row[cols["code"]])
            if code is None or code in seen_codes:
                skipped += 1
                continue
            seen_codes.add(code)
            name_ko = (row[cols["name_ko"]] or "").strip()
            name_en = (
                (row[cols["name_en"]] or "").strip()
                if "name_en" in cols
                else None
            )
            if not name_ko:
                skipped += 1
                continue
            batch.append(
                {
                    "code": code,
                    "name_ko": name_ko,
                    "name_en": name_en or None,
                }
            )
            if len(batch) >= 500:
                inserted += _flush(session, batch, dry_run=dry_run)
                batch.clear()
        if batch:
            inserted += _flush(session, batch, dry_run=dry_run)
        if not dry_run:
            session.commit()
    finally:
        session.close()

    return {"inserted_or_updated": inserted, "skipped": skipped}


def _flush(session, batch: list[dict], *, dry_run: bool) -> int:
    if dry_run:
        return len(batch)
    stmt = pg_insert(HsCode).values(batch)
    stmt = stmt.on_conflict_do_update(
        index_elements=["code"],
        set_={"name_ko": stmt.excluded.name_ko, "name_en": stmt.excluded.name_en},
    )
    session.execute(stmt)
    return len(batch)


def _main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--file", required=True, type=Path, help="Excel file path")
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and validate but don't write to DB.",
    )
    args = parser.parse_args(argv)

    result = import_file(args.file, dry_run=args.dry_run)
    print(f"inserted/updated: {result['inserted_or_updated']}")
    print(f"skipped:          {result['skipped']}")
    return 0


if __name__ == "__main__":
    sys.exit(_main())
